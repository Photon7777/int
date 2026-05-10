from __future__ import annotations

from datetime import datetime
from html import escape
import base64
import io
import json
import os
from pathlib import Path
from typing import Mapping

import altair as alt
import pandas as pd
import streamlit as st
from dotenv import load_dotenv

from marketing_mix_model import (
    CHANNEL_LABELS,
    CONTROL_VARIABLE_LABELS,
    CUSTOMER_COL,
    DATE_COL,
    DEFAULT_CONTROL_COLUMNS,
    DEFAULT_CHANNELS,
    TARGET_COL,
    build_response_curve,
    estimate_channel_contribution,
    fit_marketing_mix_model,
    generate_recommendations,
    generate_sample_marketing_data,
    get_baseline_scenario,
    normalize_marketing_data,
    optimize_budget,
    prepare_marketing_data,
    simulate_spend_change,
)

APP_ROOT = Path(__file__).resolve().parent
ASSET_DIR = APP_ROOT / "assets"


def asset_data_uri(filename: str) -> str:
    path = ASSET_DIR / filename
    if not path.exists():
        return ""
    mime = "image/svg+xml" if path.suffix.lower() == ".svg" else "image/png"
    encoded = base64.b64encode(path.read_bytes()).decode("ascii")
    return f"data:{mime};base64,{encoded}"


try:
    from marketing_mix_model import (
        DEFAULT_ADSTOCK_DECAYS,
        FIELD_LABELS,
        OPTIONAL_MODEL_COLUMNS,
        REQUIRED_MODEL_COLUMNS,
        apply_column_mapping,
        assess_data_readiness,
        available_control_columns,
        build_business_kpi_scorecard,
        build_genai_evidence_packet,
        compare_candidate_models,
        fit_bayesian_marketing_mix_model,
        predict_with_interval,
        suggest_column_mapping,
    )
except ImportError:
    DEFAULT_ADSTOCK_DECAYS = {
        "google_ads": 0.35,
        "meta_ads": 0.40,
        "instagram_ads": 0.38,
        "tv_ads": 0.65,
        "email_marketing": 0.20,
        "promotions": 0.25,
    }
    CONTROL_VARIABLE_LABELS = {
        "holiday_flag": "Holiday Flag",
        "stockout_flag": "Stockout Flag",
        "promo_event": "Promotion Event",
        "competitor_campaign": "Competitor Campaign",
        "product_launch": "Product Launch",
        "macro_index": "Macro Index",
    }
    DEFAULT_CONTROL_COLUMNS = tuple(CONTROL_VARIABLE_LABELS.keys())
    FIELD_LABELS = {DATE_COL: "Date", TARGET_COL: "Revenue", CUSTOMER_COL: "New Customers", **CHANNEL_LABELS}
    REQUIRED_MODEL_COLUMNS = (DATE_COL, TARGET_COL, *DEFAULT_CHANNELS)
    FIELD_LABELS.update(CONTROL_VARIABLE_LABELS)
    OPTIONAL_MODEL_COLUMNS = (CUSTOMER_COL, *DEFAULT_CONTROL_COLUMNS)

    def suggest_column_mapping(data: pd.DataFrame) -> dict[str, str]:
        return {column: column if column in data.columns else "" for column in (*REQUIRED_MODEL_COLUMNS, *OPTIONAL_MODEL_COLUMNS)}

    def apply_column_mapping(data: pd.DataFrame, mapping: Mapping[str, str]) -> pd.DataFrame:
        frame = data.copy()
        for canonical, source in mapping.items():
            if source and source in data.columns:
                frame[canonical] = data[source]
        return frame

    def assess_data_readiness(data: pd.DataFrame) -> dict[str, object]:
        missing = [column for column in REQUIRED_MODEL_COLUMNS if column not in normalize_marketing_data(data).columns]
        score = max(0, 100 - 12 * len(missing))
        return {
            "score": score,
            "status": "Ready" if not missing else "Needs cleanup",
            "checks": pd.DataFrame(
                [{"Area": "Required columns", "Status": "Ready" if not missing else "Needs attention", "Detail": ", ".join(missing) or "Required fields present."}]
            ),
        }

    def compare_candidate_models(*args, **kwargs) -> pd.DataFrame:
        return pd.DataFrame()

    def fit_bayesian_marketing_mix_model(data: pd.DataFrame):
        return fit_marketing_mix_model(data)

    def predict_with_interval(model, data, confidence: float = 0.80) -> pd.DataFrame:
        prediction = model.predict(data)
        error = float(model.metrics.get("rmse", 1.0))
        return pd.DataFrame({"Prediction": prediction, "Lower": prediction - error, "Upper": prediction + error})

    def available_control_columns(data: pd.DataFrame) -> tuple[str, ...]:
        frame = normalize_marketing_data(data)
        return tuple(column for column in DEFAULT_CONTROL_COLUMNS if column in frame.columns)

    def build_business_kpi_scorecard(
        data: pd.DataFrame,
        optimization: Mapping[str, object],
        model_metrics: Mapping[str, float] | None = None,
        target_roi_lift_pct: float = 5.0,
        target_cac_reduction_pct: float = 3.0,
        max_mape_pct: float = 15.0,
    ) -> pd.DataFrame:
        current_budget = float(optimization.get("current_budget", 0.0) or 0.0)
        recommended_budget = float(optimization.get("recommended_budget", current_budget) or 0.0)
        current_revenue = float(optimization.get("current_revenue", 0.0) or 0.0)
        optimized_revenue = float(optimization.get("optimized_revenue", current_revenue) or 0.0)
        current_roi = current_revenue / current_budget if current_budget else 0.0
        optimized_roi = optimized_revenue / recommended_budget if recommended_budget else 0.0
        roi_lift = ((optimized_roi - current_roi) / current_roi * 100) if current_roi else 0.0
        mape = float((model_metrics or {}).get("mape", 0.0))
        return pd.DataFrame(
            [
                {
                    "Business KPI": "Marketing ROI lift",
                    "Current Value": current_roi,
                    "Projected Value": optimized_roi,
                    "Delta %": roi_lift,
                    "Target": target_roi_lift_pct,
                    "Status": "Met" if roi_lift >= target_roi_lift_pct else "Watch",
                    "Interpretation": "Revenue generated per marketing dollar after budget reallocation.",
                },
                {
                    "Business KPI": "CAC reduction",
                    "Current Value": None,
                    "Projected Value": None,
                    "Delta %": None,
                    "Target": target_cac_reduction_pct,
                    "Status": "Needs customer data",
                    "Interpretation": "Upload customer or conversion counts to quantify CAC impact.",
                },
                {
                    "Business KPI": "Forecast quality",
                    "Current Value": mape,
                    "Projected Value": mape,
                    "Delta %": None,
                    "Target": max_mape_pct,
                    "Status": "Met" if mape <= max_mape_pct else "Watch",
                    "Interpretation": "Lower MAPE means the ML layer is reliable enough for budget planning.",
                },
            ]
        )

    def build_genai_evidence_packet(*args, **kwargs) -> dict[str, object]:
        return {
            "selected_ai_approach": "Both: predictive ML plus grounded generative AI",
            "generation_layer": {"strategy": "Ground recommendations in MMM outputs."},
        }

try:
    from marketing_mix_model import evaluate_model_against_baseline
except ImportError:
    def evaluate_model_against_baseline(*args, **kwargs):
        raise ValueError("Train/test evaluation requires the latest marketing_mix_model.py.")


load_dotenv()

st.set_page_config(
    page_title="Mixalyzer",
    page_icon="M",
    layout="wide",
    initial_sidebar_state="expanded",
)


st.markdown(
    """
    <style>
      #MainMenu, footer {visibility: hidden;}
      .stApp {
        background:
          radial-gradient(circle at 12% 0%, rgba(56, 215, 193, 0.11), transparent 34%),
          radial-gradient(circle at 90% 6%, rgba(246, 200, 95, 0.08), transparent 30%),
          linear-gradient(180deg, #07111a 0%, #0b1420 48%, #0d1117 100%);
      }
      section[data-testid="stSidebar"] {
        background: linear-gradient(180deg, #0B1424 0%, #111827 100%);
        border-right: 1px solid rgba(148, 163, 184, 0.16);
      }
      .block-container {
        max-width: 1420px;
        padding-top: 1.4rem;
        padding-bottom: 2.5rem;
      }
      .app-header {
        display: grid;
        grid-template-columns: minmax(0, 1fr) auto;
        gap: 1.25rem;
        align-items: start;
        margin-bottom: 1rem;
      }
      .brand-lockup {
        display: flex;
        flex-direction: column;
        align-items: flex-start;
        gap: 0.85rem;
      }
      .brand-logo {
        width: min(320px, 100%);
        height: auto;
        display: block;
      }
      .brand-copy {
        max-width: 960px;
      }
      .brand-mark {
        width: 54px;
        height: 54px;
        border-radius: 16px;
        background: linear-gradient(135deg, #38D7C1, #7C8CFF 58%, #F6C85F);
        display: grid;
        place-items: center;
        color: #07111A;
        font-weight: 900;
        font-size: 1.1rem;
        box-shadow: 0 16px 42px rgba(15, 23, 42, 0.34);
      }
      .brand-kicker {
        color: #24c6a1;
        font-size: 0.84rem;
        font-weight: 800;
        letter-spacing: 0.08em;
        text-transform: uppercase;
        margin-bottom: 0.2rem;
      }
      .brand-title {
        color: #f8fafc;
        font-size: 2.65rem;
        font-weight: 850;
        line-height: 1.06;
        margin: 0 0 0.35rem;
      }
      .brand-subtitle {
        color: rgba(226, 232, 240, 0.76);
        max-width: 960px;
        font-size: 1rem;
        line-height: 1.5;
        margin-bottom: 0.4rem;
      }
      .brand-pill {
        border: 1px solid rgba(56, 215, 193, 0.32);
        background: rgba(56, 215, 193, 0.08);
        color: #A7F3D0;
        border-radius: 999px;
        padding: 0.42rem 0.78rem;
        font-weight: 800;
        font-size: 0.86rem;
        white-space: nowrap;
        margin-top: 0.45rem;
      }
      .metric-label {
        color: rgba(226, 232, 240, 0.72);
        font-size: 0.82rem;
        margin-bottom: 0.25rem;
      }
      .metric-value {
        color: #f8fafc;
        font-size: 1.65rem;
        font-weight: 750;
        line-height: 1.1;
      }
      .metric-delta {
        color: #38d7c1;
        font-size: 0.82rem;
        margin-top: 0.18rem;
      }
      .kpi-card {
        border: 1px solid rgba(148, 163, 184, 0.18);
        background: linear-gradient(180deg, rgba(19, 29, 38, 0.82), rgba(12, 19, 27, 0.84));
        border-radius: 8px;
        padding: 0.8rem 0.95rem;
        min-height: 104px;
        display: flex;
        flex-direction: column;
        justify-content: center;
        overflow: visible;
      }
      .kpi-label {
        color: rgba(226, 232, 240, 0.82);
        font-size: 0.86rem;
        font-weight: 650;
        line-height: 1.2;
        margin-bottom: 0.55rem;
        white-space: normal;
        overflow-wrap: anywhere;
      }
      .kpi-value {
        color: #f8fafc;
        font-size: 1.55rem;
        font-weight: 720;
        line-height: 1.16;
        letter-spacing: 0;
        white-space: normal;
        overflow-wrap: anywhere;
        word-break: normal;
      }
      .kpi-delta {
        color: #38d7c1;
        font-size: 0.86rem;
        font-weight: 650;
        line-height: 1.2;
        margin-top: 0.45rem;
        white-space: normal;
        overflow-wrap: anywhere;
      }
      .panel {
        border: 1px solid rgba(148, 163, 184, 0.18);
        background: rgba(13, 21, 29, 0.76);
        border-radius: 8px;
        padding: 1rem;
      }
      .recommendation {
        border-left: 3px solid #38d7c1;
        background: rgba(56, 215, 193, 0.08);
        padding: 0.7rem 0.8rem;
        margin-bottom: 0.65rem;
        border-radius: 6px;
        color: #e5eefb;
      }
      .ai-brief {
        border: 1px solid rgba(124, 140, 255, 0.28);
        background: linear-gradient(135deg, rgba(124, 140, 255, 0.13), rgba(11, 18, 32, 0.8));
        padding: 0.95rem 1rem;
        margin-bottom: 0.8rem;
        border-radius: 8px;
        color: #e5eefb;
      }
      .ai-brief h4 {
        color: #f8fafc;
        margin: 0 0 0.4rem;
      }
      .ai-brief p {
        color: rgba(226, 232, 240, 0.82);
        margin: 0.28rem 0;
      }
      .summary-box {
        border: 1px solid rgba(56, 215, 193, 0.34);
        background: linear-gradient(135deg, rgba(20, 184, 166, 0.14), rgba(15, 23, 42, 0.78));
        border-radius: 8px;
        padding: 1rem 1.1rem;
        margin: 0.75rem 0 1rem;
      }
      .summary-box h4 {
        color: #f8fafc;
        margin: 0 0 0.45rem;
      }
      .summary-box p {
        color: rgba(226, 232, 240, 0.82);
        margin: 0.18rem 0;
      }
      .badge {
        display: inline-block;
        border-radius: 999px;
        padding: 0.28rem 0.65rem;
        font-size: 0.82rem;
        font-weight: 750;
        margin-bottom: 0.4rem;
      }
      .badge-success {
        background: rgba(36, 198, 161, 0.16);
        color: #5eead4;
        border: 1px solid rgba(45, 212, 191, 0.4);
      }
      .badge-warning {
        background: rgba(245, 158, 11, 0.15);
        color: #fbbf24;
        border: 1px solid rgba(245, 158, 11, 0.4);
      }
      .badge-danger {
        background: rgba(251, 113, 133, 0.15);
        color: #fb7185;
        border: 1px solid rgba(251, 113, 133, 0.4);
      }
      .badge-info {
        background: rgba(148, 163, 184, 0.14);
        color: #cbd5e1;
        border: 1px solid rgba(148, 163, 184, 0.34);
      }
      .hero {
        border: 1px solid rgba(148, 163, 184, 0.18);
        background:
          linear-gradient(135deg, rgba(36, 198, 161, 0.14), rgba(124, 140, 255, 0.10), rgba(246, 200, 95, 0.06)),
          rgba(13, 21, 29, 0.82);
        border-radius: 8px;
        padding: 1.1rem;
        margin-bottom: 1rem;
        display: grid;
        grid-template-columns: minmax(0, 0.95fr) minmax(300px, 1.05fr);
        gap: 1.35rem;
        align-items: center;
      }
      .hero h2 {
        color: #f8fafc;
        font-size: 2.35rem;
        line-height: 1.12;
        margin: 0 0 0.6rem;
      }
      .hero p {
        color: rgba(226, 232, 240, 0.78);
        font-size: 1rem;
        margin: 0 0 0.8rem;
        line-height: 1.55;
      }
      .hero-eyebrow {
        color: #38D7C1;
        font-size: 0.78rem;
        font-weight: 850;
        letter-spacing: 0.08em;
        text-transform: uppercase;
        margin-bottom: 0.45rem;
      }
      .hero-visual {
        width: 100%;
        border-radius: 8px;
        display: block;
        box-shadow: 0 28px 74px rgba(2, 6, 23, 0.28);
      }
      .identity-row {
        display: flex;
        gap: 0.55rem;
        flex-wrap: wrap;
        margin-top: 0.8rem;
      }
      .identity-chip {
        border: 1px solid rgba(148, 163, 184, 0.22);
        background: rgba(15, 23, 42, 0.55);
        color: #dbeafe;
        border-radius: 999px;
        padding: 0.38rem 0.62rem;
        font-size: 0.84rem;
        font-weight: 700;
      }
      .feature-card {
        border: 1px solid rgba(148, 163, 184, 0.16);
        background: linear-gradient(180deg, rgba(15, 23, 42, 0.76), rgba(13, 21, 29, 0.68));
        border-radius: 8px;
        padding: 0.9rem;
        min-height: 120px;
      }
      .feature-card h4 {
        color: #f8fafc;
        margin: 0 0 0.35rem;
      }
      .feature-card p {
        color: rgba(226, 232, 240, 0.74);
        margin: 0;
      }
      .risk-ok {
        border-left: 3px solid #38d7c1;
      }
      .risk-watch {
        border-left: 3px solid #f59e0b;
      }
      .readiness-good {
        border-left: 4px solid #24c6a1;
      }
      .readiness-watch {
        border-left: 4px solid #f6c85f;
      }
      .readiness-bad {
        border-left: 4px solid #fb7185;
      }
      div[data-testid="stTabs"] button p {
        font-weight: 700;
      }
      div[data-testid="stMetric"] {
        border: 1px solid rgba(148, 163, 184, 0.16);
        background: rgba(15, 23, 42, 0.55);
        border-radius: 8px;
        padding: 0.8rem 0.95rem;
      }
      div[data-testid="stMetricValue"],
      div[data-testid="stMetricValue"] div {
        white-space: normal !important;
        overflow: visible !important;
        text-overflow: clip !important;
      }
      @media (max-width: 900px) {
        .app-header,
        .hero {
          grid-template-columns: 1fr;
        }
        .brand-pill {
          justify-self: start;
          margin-top: 0;
        }
      }
    </style>
    """,
    unsafe_allow_html=True,
)


@st.cache_data(show_spinner=False)
def load_sample_data() -> pd.DataFrame:
    return generate_sample_marketing_data()


@st.cache_data(show_spinner=False)
def train_model(data: pd.DataFrame, regularization: float, model_engine: str):
    if model_engine.startswith("Bayesian"):
        model = fit_bayesian_marketing_mix_model(data)
    else:
        model = fit_marketing_mix_model(data, regularization=regularization)
    contribution = estimate_channel_contribution(model, data)
    baseline = get_baseline_scenario(data)
    return model, contribution, baseline


@st.cache_data(show_spinner=False)
def evaluate_current_model(data: pd.DataFrame, regularization: float):
    return evaluate_model_against_baseline(data, regularization=regularization)


@st.cache_data(show_spinner=False)
def compare_models(data: pd.DataFrame, regularization: float):
    return compare_candidate_models(data, regularization=regularization)


def money(value: float) -> str:
    return f"${value:,.0f}"


def pct(value: float) -> str:
    return f"{value:.1f}%"


def signed_money(value: float) -> str:
    prefix = "+" if value >= 0 else "-"
    return f"{prefix}${abs(value):,.0f}"


def scorecard_value(metric: str, value: object) -> str:
    if pd.isna(value):
        return "N/A"
    numeric = float(value)
    if metric == "Forecast quality":
        return pct(numeric)
    if "CAC" in metric:
        return money(numeric)
    return f"{numeric:.2f}x"


def scorecard_delta(value: object) -> str:
    if pd.isna(value):
        return "N/A"
    return pct(float(value))


def display_business_scorecard(scorecard: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for _, row in scorecard.iterrows():
        rows.append(
            {
                "Business KPI": row["Business KPI"],
                "Current": scorecard_value(row["Business KPI"], row["Current Value"]),
                "Projected": scorecard_value(row["Business KPI"], row["Projected Value"]),
                "Delta": scorecard_delta(row["Delta %"]),
                "Target": pct(float(row["Target"])),
                "Status": row["Status"],
                "Interpretation": row["Interpretation"],
            }
        )
    return pd.DataFrame(rows)


def model_trust_assessment(evaluation: Mapping[str, object] | None, max_mape_pct: float) -> dict[str, object]:
    if not evaluation:
        return {
            "label": "Insufficient data",
            "status": "info",
            "mmm_beats_baseline": False,
            "mape": None,
            "baseline_mape": None,
            "explanation": "There is not enough holdout data to compare MMM against a baseline.",
            "next_step": "Improve data coverage before relying on optimization.",
        }

    model_mape = float(evaluation["model_metrics"]["mape"])
    baseline_mape = float(evaluation["baseline_metrics"]["mape"])
    mmm_beats_baseline = model_mape < baseline_mape

    if mmm_beats_baseline and model_mape <= float(max_mape_pct):
        label = "Trust for planning"
        status = "success"
        next_step = "Pilot the recommendation and monitor weekly."
    elif mmm_beats_baseline:
        label = "Use with caution"
        status = "warning"
        next_step = "Use a limited pilot and improve data quality before larger budget changes."
    else:
        label = "Do not use recommendation yet"
        status = "danger"
        next_step = "Improve the model or dataset before acting on optimization."

    explanation = (
        f"MMM MAPE is {model_mape:.1f}% vs baseline MAPE of {baseline_mape:.1f}%. "
        f"Your planning target is {float(max_mape_pct):.1f}%."
    )
    return {
        "label": label,
        "status": status,
        "mmm_beats_baseline": mmm_beats_baseline,
        "mape": model_mape,
        "baseline_mape": baseline_mape,
        "explanation": explanation,
        "next_step": next_step,
    }


def render_badge(label: str, status: str = "info") -> None:
    st.markdown(
        f"<span class='badge badge-{escape(status)}'>{escape(label)}</span>",
        unsafe_allow_html=True,
    )


def scorecard_metric(scorecard: pd.DataFrame, metric: str, column: str) -> float | None:
    if scorecard.empty or metric not in set(scorecard["Business KPI"]):
        return None
    value = scorecard.loc[scorecard["Business KPI"] == metric, column].iloc[0]
    return None if pd.isna(value) else float(value)


def allocation_shift_summary(optimization: Mapping[str, object]) -> dict[str, object]:
    allocation = optimization.get("allocation")
    if not isinstance(allocation, pd.DataFrame) or allocation.empty:
        return {"increase": "higher-response channels", "decrease": "lower-ROI channels", "increase_rows": pd.DataFrame(), "decrease_rows": pd.DataFrame()}
    increase_rows = allocation[allocation["Spend Shift"] > 0].sort_values("Spend Shift", ascending=False)
    decrease_rows = allocation[allocation["Spend Shift"] < 0].sort_values("Spend Shift", ascending=True)
    return {
        "increase": ", ".join(increase_rows["Channel"].head(2).astype(str)) or "higher-response channels",
        "decrease": ", ".join(decrease_rows["Channel"].head(2).astype(str)) or "lower-ROI channels",
        "increase_rows": increase_rows,
        "decrease_rows": decrease_rows,
    }


def executive_recommendation_summary(
    optimization: Mapping[str, object],
    scorecard: pd.DataFrame,
    evaluation: Mapping[str, object] | None,
    max_mape_pct: float,
) -> dict[str, str]:
    trust = model_trust_assessment(evaluation, max_mape_pct)
    shifts = allocation_shift_summary(optimization)
    roi_lift = scorecard_metric(scorecard, "Marketing ROI lift", "Delta %")
    cac_lift = scorecard_metric(scorecard, "CAC reduction", "Delta %")
    confidence = (
        "Conservative case remains positive."
        if float(optimization.get("revenue_delta_low", 0.0) or 0.0) > 0
        else "Conservative case is not yet positive."
    )
    if cac_lift is None:
        cac_phrase = "CAC unavailable without customer/conversion data"
    elif cac_lift >= 0:
        cac_phrase = f"{cac_lift:+.1f}% CAC reduction"
    else:
        cac_phrase = f"{abs(cac_lift):.1f}% CAC increase risk"
    return {
        "recommendation": (
            f"Reallocate part of next period's budget from {shifts['decrease']} into {shifts['increase']}."
        ),
        "impact": (
            f"Expected impact: {float(optimization.get('revenue_delta_pct', 0.0) or 0.0):+.1f}% revenue lift, "
            f"{roi_lift:+.1f}% ROI lift, {cac_phrase}."
            if roi_lift is not None
            else f"Expected impact: {float(optimization.get('revenue_delta_pct', 0.0) or 0.0):+.1f}% revenue lift."
        ),
        "confidence": (
            f"{confidence} Range: {signed_money(float(optimization.get('revenue_delta_low', 0.0) or 0.0))} "
            f"to {signed_money(float(optimization.get('revenue_delta_high', 0.0) or 0.0))}."
        ),
        "risk": f"Model trust: {trust['label']}. {trust['explanation']}",
        "next_step": trust["next_step"],
    }


def render_executive_summary(summary: Mapping[str, str]) -> None:
    st.markdown(
        f"""
        <div class="summary-box">
          <h4>Executive recommendation summary</h4>
          <p><strong>Recommendation:</strong> {escape(summary['recommendation'])}</p>
          <p><strong>Expected impact:</strong> {escape(summary['impact'])}</p>
          <p><strong>Confidence:</strong> {escape(summary['confidence'])}</p>
          <p><strong>Risk:</strong> {escape(summary['risk'])}</p>
          <p><strong>Recommended next step:</strong> {escape(summary['next_step'])}</p>
        </div>
        """,
        unsafe_allow_html=True,
    )


def detected_control_summary(data: pd.DataFrame) -> pd.DataFrame:
    controls = available_control_columns(data)
    if not controls:
        return pd.DataFrame(
            [
                {
                    "Control Variable": "Optional controls not detected",
                    "Status": "Missing",
                    "Why It Matters": "Holidays, stockouts, product launches, competitor campaigns, and macro indicators can improve model reliability.",
                }
            ]
        )
    return pd.DataFrame(
        [
            {
                "Control Variable": CONTROL_VARIABLE_LABELS.get(column, column),
                "Status": "Included in model",
                "Why It Matters": "Helps prevent marketing channels from receiving credit or blame for external business events.",
            }
            for column in controls
        ]
    )


def build_pilot_plan(
    optimization: Mapping[str, object],
    scorecard: pd.DataFrame,
    evaluation: Mapping[str, object] | None,
    max_mape_pct: float,
    duration_weeks: int = 4,
) -> pd.DataFrame:
    shifts = allocation_shift_summary(optimization)
    trust = model_trust_assessment(evaluation, max_mape_pct)
    cac_delta = scorecard_metric(scorecard, "CAC reduction", "Delta %")
    if cac_delta is None:
        cac_guardrail = "if CAC rises beyond target"
    elif cac_delta >= 0:
        cac_guardrail = f"if CAC reduction falls below target ({cac_delta:.1f}% current estimate)"
    else:
        cac_guardrail = f"if CAC increases beyond target ({abs(cac_delta):.1f}% current estimate)"
    return pd.DataFrame(
        [
            {
                "Plan Area": "Recommended budget shift",
                "Action": f"Move spend from {shifts['decrease']} into {shifts['increase']}.",
            },
            {"Plan Area": "Pilot duration", "Action": f"Run a monitored {duration_weeks}-week pilot before permanent rollout."},
            {
                "Plan Area": "Channels to increase",
                "Action": shifts["increase"],
            },
            {
                "Plan Area": "Channels to decrease",
                "Action": shifts["decrease"],
            },
            {
                "Plan Area": "KPIs to monitor",
                "Action": "Revenue, marketing ROI, CAC, MAPE, and conversion volume.",
            },
            {"Plan Area": "Monitoring cadence", "Action": "Review performance weekly with marketing and finance."},
            {
                "Plan Area": "Stop-loss rule",
                "Action": (
                    "Pause or revise if conservative lift is negative for 2 consecutive weeks, "
                    f"{cac_guardrail}, or if model error exceeds {float(max_mape_pct):.1f}% MAPE."
                ),
            },
            {
                "Plan Area": "Human review checkpoint",
                "Action": f"Before permanent rollout, confirm business context and model trust status: {trust['label']}.",
            },
        ]
    )


def render_metric_card(label: object, value: object, delta: object | None = None) -> None:
    delta_markup = (
        f'<div class="kpi-delta">{escape(str(delta))}</div>'
        if delta is not None and str(delta) != ""
        else ""
    )
    st.markdown(
        f"""
        <div class="kpi-card">
          <div class="kpi-label">{escape(str(label))}</div>
          <div class="kpi-value">{escape(str(value))}</div>
          {delta_markup}
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_loader(message: str) -> None:
    st.markdown(
        f"""
        <div class="panel">
          <strong>{escape(message)}</strong>
        </div>
        """,
        unsafe_allow_html=True,
    )


def dataframe_to_markdown_table(frame: pd.DataFrame) -> str:
    headers = list(frame.columns)
    rows = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join(["---"] * len(headers)) + " |",
    ]
    for _, row in frame.iterrows():
        cells = []
        for value in row:
            if isinstance(value, float):
                cells.append(f"{value:,.2f}")
            else:
                cells.append(str(value))
        rows.append("| " + " | ".join(cells) + " |")
    return "\n".join(rows)


def read_input_data(upload) -> pd.DataFrame:
    if upload is None:
        return load_sample_data()
    return pd.read_csv(upload)


def build_column_mapping_controls(raw_frame: pd.DataFrame, upload_present: bool) -> dict[str, str]:
    if not upload_present:
        return {column: column for column in (*REQUIRED_MODEL_COLUMNS, *OPTIONAL_MODEL_COLUMNS)}

    suggestions = suggest_column_mapping(raw_frame)
    options = [""] + list(raw_frame.columns)
    mapping: dict[str, str] = {}
    for canonical in (*REQUIRED_MODEL_COLUMNS, *OPTIONAL_MODEL_COLUMNS):
        suggested = suggestions.get(canonical, "")
        default_index = options.index(suggested) if suggested in options else 0
        mapping[canonical] = st.selectbox(
            FIELD_LABELS.get(canonical, canonical),
            options=options,
            index=default_index,
            key=f"column_map_{canonical}",
        )
    return mapping


def readiness_class(status: str) -> str:
    if status == "Ready":
        return "readiness-good"
    if status == "Usable with caveats":
        return "readiness-watch"
    return "readiness-bad"


def confidence_range_chart(summary: pd.DataFrame) -> alt.Chart:
    return (
        alt.Chart(summary)
        .mark_bar(cornerRadius=5)
        .encode(
            x=alt.X("Case:N", title=None),
            y=alt.Y("Revenue Delta:Q", title="Predicted revenue impact"),
            color=alt.Color(
                "Case:N",
                scale=alt.Scale(range=["#fb7185", "#24c6a1", "#f6c85f"]),
                legend=None,
            ),
            tooltip=[
                alt.Tooltip("Case:N"),
                alt.Tooltip("Revenue Delta:Q", format="$,.0f"),
            ],
        )
        .properties(height=240)
    )


def model_comparison_chart(comparison: pd.DataFrame) -> alt.Chart:
    return (
        alt.Chart(comparison)
        .mark_bar(cornerRadiusTopLeft=4, cornerRadiusTopRight=4)
        .encode(
            x=alt.X("Model:N", sort=alt.SortField("MAPE", order="ascending"), title=None),
            y=alt.Y("MAPE:Q", title="MAPE"),
            color=alt.Color(
                "Model:N",
                scale=alt.Scale(range=["#24c6a1", "#8ab4ff", "#f6c85f", "#fb7185"]),
                legend=None,
            ),
            tooltip=[
                alt.Tooltip("Model:N"),
                alt.Tooltip("MAPE:Q", format=".2f"),
                alt.Tooltip("RMSE:Q", format="$,.0f"),
                alt.Tooltip("R2:Q", format=".2f"),
            ],
        )
        .properties(height=300)
    )


def simulate_with_confidence(model, baseline: Mapping[str, object], changes: Mapping[str, float], confidence: float):
    try:
        return simulate_spend_change(model, baseline, changes, confidence=confidence)
    except TypeError:
        result = simulate_spend_change(model, baseline, changes)
        spread = float(model.metrics.get("rmse", 0.0) or 0.0)
        result.setdefault("revenue_delta_low", result["revenue_delta"] - spread)
        result.setdefault("revenue_delta_high", result["revenue_delta"] + spread)
        return result


def optimize_with_confidence(model, baseline: Mapping[str, object], budget: float, confidence: float):
    try:
        return optimize_budget(model, baseline, total_budget=budget, confidence=confidence)
    except TypeError:
        result = optimize_budget(model, baseline, total_budget=budget)
        spread = float(model.metrics.get("rmse", 0.0) or 0.0)
        result.setdefault("revenue_delta_low", result["revenue_delta"] - spread)
        result.setdefault("revenue_delta_high", result["revenue_delta"] + spread)
        return result


def build_csv_template() -> bytes:
    template = pd.DataFrame(
        [
            {
                DATE_COL: "2026-01-04",
                "google_ads": 12000,
                "meta_ads": 9000,
                "instagram_ads": 7000,
                "tv_ads": 18000,
                "email_marketing": 2500,
                "promotions": 4500,
                "holiday_flag": 0,
                "stockout_flag": 0,
                "promo_event": 1,
                "competitor_campaign": 0,
                "product_launch": 0,
                "macro_index": 101.2,
                TARGET_COL: 260000,
                CUSTOMER_COL: 720,
            }
        ]
    )
    return template.to_csv(index=False).encode("utf-8")


def prediction_chart(predictions: pd.DataFrame) -> alt.Chart:
    long = predictions.melt(
        id_vars=[DATE_COL],
        value_vars=["Actual Revenue", "MMX Prediction", "Baseline Prediction"],
        var_name="Series",
        value_name="Revenue",
    )
    return (
        alt.Chart(long)
        .mark_line(point=True, strokeWidth=3)
        .encode(
            x=alt.X(f"{DATE_COL}:T", title=None),
            y=alt.Y("Revenue:Q", title="Revenue"),
            color=alt.Color(
                "Series:N",
                scale=alt.Scale(range=["#f8fafc", "#38d7c1", "#f97316"]),
                legend=alt.Legend(orient="top"),
            ),
            tooltip=[
                alt.Tooltip(f"{DATE_COL}:T", title="Week"),
                alt.Tooltip("Series:N"),
                alt.Tooltip("Revenue:Q", format="$,.0f"),
            ],
        )
        .properties(height=330)
    )


def evaluation_metric_chart(evaluation: Mapping[str, object]) -> alt.Chart:
    model_metrics = evaluation["model_metrics"]
    baseline_metrics = evaluation["baseline_metrics"]
    rows = [
        {"Metric": "MAPE", "Model": "MMX", "Value": model_metrics["mape"]},
        {"Metric": "MAPE", "Model": "Baseline", "Value": baseline_metrics["mape"]},
        {"Metric": "RMSE", "Model": "MMX", "Value": model_metrics["rmse"]},
        {"Metric": "RMSE", "Model": "Baseline", "Value": baseline_metrics["rmse"]},
    ]
    return (
        alt.Chart(pd.DataFrame(rows))
        .mark_bar(cornerRadiusTopLeft=4, cornerRadiusTopRight=4)
        .encode(
            x=alt.X("Metric:N", title=None),
            y=alt.Y("Value:Q", title=None),
            color=alt.Color(
                "Model:N",
                scale=alt.Scale(range=["#38d7c1", "#f97316"]),
                legend=alt.Legend(orient="top"),
            ),
            xOffset="Model:N",
            tooltip=[alt.Tooltip("Model:N"), alt.Tooltip("Metric:N"), alt.Tooltip("Value:Q", format=",.2f")],
        )
        .properties(height=300)
    )


def build_responsible_ai_audit(
    data: pd.DataFrame,
    evaluation: Mapping[str, object] | None,
    contribution: pd.DataFrame,
    max_mape_pct: float = 15.0,
) -> pd.DataFrame:
    weeks = len(data)
    normalized_columns = {column_name.lower() for column_name in map(str, data.columns)}
    privacy_terms = {"email", "phone", "name", "customer_id", "user_id"}
    privacy_hits = sorted(
        column for column in normalized_columns if any(term in column for term in privacy_terms)
    )
    segment_terms = {"region", "segment", "market", "geo", "product", "audience"}
    segment_hits = sorted(
        column for column in normalized_columns if any(term in column for term in segment_terms)
    )
    trust = model_trust_assessment(evaluation, max_mape_pct)
    readiness_issues = []
    if weeks < 52:
        readiness_issues.append("short history")
    if DATE_COL in data.columns:
        dates = pd.to_datetime(data[DATE_COL], errors="coerce").sort_values()
        if dates.isna().any():
            readiness_issues.append("invalid dates")
        elif len(dates) > 2 and dates.diff().dropna().dt.days.nunique() > 2:
            readiness_issues.append("irregular dates")
    numeric_columns = [column for column in DEFAULT_CHANNELS if column in data.columns]
    for column in numeric_columns:
        numeric = pd.to_numeric(data[column], errors="coerce")
        if numeric.isna().any():
            readiness_issues.append(f"non-numeric {column}")
            break

    return pd.DataFrame(
        [
        {
            "Risk Area": "Privacy",
            "Example in Mixalyzer": (
                "Uploaded data may contain customer identifiers: " + ", ".join(privacy_hits)
                if privacy_hits
                else "Uploaded data appears aggregated with no obvious customer identifiers."
            ),
            "Potential Impact": "Sensitive identifiers could be exposed in analysis files or exports.",
            "Mitigation / Guardrail": "Use weekly aggregated data only; remove email, phone, name, customer_id, and user_id columns before sharing.",
            "Status": "Needs review" if privacy_hits else "Passed",
        },
        {
            "Risk Area": "Bias / fairness",
            "Example in Mixalyzer": (
                "Segment fields detected: " + ", ".join(segment_hits)
                if segment_hits
                else "No region, product, audience, or customer segment fields detected."
            ),
            "Potential Impact": "The model may overrepresent channels that served only one region, product, or customer segment.",
            "Mitigation / Guardrail": "Review results by segment when segment fields exist; avoid making broad cuts from one aggregate model alone.",
            "Status": "Warning" if segment_hits else "Warning",
        },
        {
            "Risk Area": "Reliability",
            "Example in Mixalyzer": trust["explanation"],
            "Potential Impact": "Weak model accuracy can lead to wrong budget reallocations.",
            "Mitigation / Guardrail": "Compare MMM against baselines, use the model trust badge, and avoid optimization if baseline performs better.",
            "Status": "Passed" if trust["status"] == "success" else "Needs review" if trust["status"] == "danger" else "Warning",
        },
        {
            "Risk Area": "Hallucination",
            "Example in Mixalyzer": "AI-generated recommendations may invent unsupported channel claims.",
            "Potential Impact": "Stakeholders could act on a persuasive but unsupported narrative.",
            "Mitigation / Guardrail": "Generate recommendations only from the evidence workbook and model outputs; fall back to deterministic recommendations.",
            "Status": "Passed",
        },
        {
            "Risk Area": "Over-automation",
            "Example in Mixalyzer": "User applies budget changes without business review.",
            "Potential Impact": "Budget shifts could harm revenue, CAC, or strategic brand goals.",
            "Mitigation / Guardrail": "Show decision-support warning, recommend a pilot rollout, and require a human checkpoint before permanent rollout.",
            "Status": "Warning",
        },
        {
            "Risk Area": "Data quality",
            "Example in Mixalyzer": (
                "Readiness risks detected: " + ", ".join(readiness_issues)
                if readiness_issues
                else "Required columns are present and no obvious date/numeric issues were detected."
            ),
            "Potential Impact": "Missing weeks, irregular dates, short history, or non-numeric spend can distort ROI and contribution.",
            "Mitigation / Guardrail": "Use readiness checks before model training and improve the dataset when warnings appear.",
            "Status": "Warning" if readiness_issues else "Passed",
        },
        ]
    )


def build_executive_report(
    data: pd.DataFrame,
    contribution: pd.DataFrame,
    optimization: Mapping[str, object],
    simulation: Mapping[str, object],
    evaluation: Mapping[str, object] | None,
    recommendations: list[str],
    scorecard: pd.DataFrame | None = None,
) -> str:
    total_spend_value = float(data[list(DEFAULT_CHANNELS)].sum().sum())
    total_revenue_value = float(data[TARGET_COL].sum())
    total_customers_value = float(data[CUSTOMER_COL].sum()) if CUSTOMER_COL in data.columns else 0.0
    cac_value = total_spend_value / total_customers_value if total_customers_value else None
    top_roi = contribution.sort_values("ROI", ascending=False).iloc[0]
    weakest_roi = contribution.sort_values("ROI", ascending=True).iloc[0]
    allocation = optimization["allocation"][
        ["Channel", "Current Spend", "Recommended Spend", "Spend Shift", "Change %"]
    ].copy()

    lines = [
        "# Marketing Mix Optimization Executive Report",
        "",
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        "",
        "## Business KPI Snapshot",
        f"- Revenue analyzed: {money(total_revenue_value)}",
        f"- Marketing spend analyzed: {money(total_spend_value)}",
        f"- Marketing ROI: {total_revenue_value / total_spend_value:.2f}x" if total_spend_value else "- Marketing ROI: N/A",
        f"- CAC: {money(cac_value)}" if cac_value is not None else "- CAC: N/A",
        "",
        "## Recommended Budget Action",
        f"- Current next-period budget: {money(optimization['current_budget'])}",
        f"- Recommended next-period budget: {money(optimization['recommended_budget'])}",
        f"- Predicted revenue impact: {signed_money(optimization['revenue_delta'])} ({optimization['revenue_delta_pct']:.1f}%)",
        "",
        "## Business Goal Attainment",
    ]

    if scorecard is not None and not scorecard.empty:
        for _, row in display_business_scorecard(scorecard).iterrows():
            lines.append(
                f"- {row['Business KPI']}: {row['Status']} "
                f"(current {row['Current']}, projected {row['Projected']}, delta {row['Delta']}, target {row['Target']})"
            )
    else:
        lines.append("- Business KPI scorecard unavailable.")

    lines.extend(
        [
            "",
            "## Channel Insights",
            f"- Highest estimated ROI: {top_roi['Channel']} at {float(top_roi['ROI']):.2f}x",
            f"- Lowest estimated ROI: {weakest_roi['Channel']} at {float(weakest_roi['ROI']):.2f}x",
            "",
            "## Model Evaluation",
        ]
    )

    if evaluation:
        lines.extend(
            [
                f"- Train rows: {evaluation['train_rows']}; test rows: {evaluation['test_rows']}",
                f"- MMX MAPE: {evaluation['model_metrics']['mape']:.2f}%",
                f"- Baseline MAPE: {evaluation['baseline_metrics']['mape']:.2f}%",
                f"- RMSE improvement vs baseline: {evaluation['rmse_improvement_pct']:.1f}%",
            ]
        )
    else:
        lines.append("- Evaluation unavailable because the uploaded dataset is too short.")

    lines.extend(["", "## Active Simulation", f"- Predicted revenue impact: {signed_money(simulation['revenue_delta'])} ({simulation['revenue_delta_pct']:.1f}%)", ""])
    lines.append("## Recommended Allocation")
    lines.append(dataframe_to_markdown_table(allocation))
    lines.extend(["", "## Management Recommendations"])
    lines.extend([f"- {item}" for item in recommendations])
    lines.extend(
        [
            "",
            "## Responsible AI Notes",
            "- Treat recommendations as decision support, not automated media-buying instructions.",
            "- Validate large reallocations with incrementality tests or controlled experiments.",
            "- Use aggregated spend/revenue/customer data and avoid customer-level identifiers.",
        ]
    )
    return "\n".join(lines)


def _clean_export_value(value: object) -> object:
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, default=str)
    if pd.isna(value):
        return ""
    return value


def _style_workbook(workbook) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="0F766E")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(bottom=Side(style="thin", color="CBD5E1"))

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        for column_cells in worksheet.columns:
            values = [str(cell.value or "") for cell in column_cells]
            width = min(max(max(len(value) for value in values) + 2, 12), 48)
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width


def _write_excel_frame(writer: pd.ExcelWriter, frame: pd.DataFrame, sheet_name: str) -> None:
    cleaned = frame.copy()
    for column in cleaned.columns:
        cleaned[column] = cleaned[column].map(_clean_export_value)
    cleaned.to_excel(writer, sheet_name=sheet_name[:31], index=False)


def _evidence_packet_sheets(evidence_packet: Mapping[str, object]) -> dict[str, pd.DataFrame]:
    prediction = evidence_packet.get("prediction_layer", {}) or {}
    generation = evidence_packet.get("generation_layer", {}) or {}
    optimization = evidence_packet.get("optimization", {}) or {}
    targets = evidence_packet.get("business_targets", {}) or {}

    prediction_rows = [
        {"Item": "Selected AI approach", "Value": evidence_packet.get("selected_ai_approach", "")},
        {"Item": "Prediction model", "Value": prediction.get("model", "")},
        {"Item": "Model features", "Value": ", ".join(prediction.get("features", []))},
    ]
    prediction_rows.extend(
        {"Item": f"Metric: {metric}", "Value": value}
        for metric, value in (prediction.get("metrics", {}) or {}).items()
    )

    generation_rows = [
        {"Item": "Strategy", "Value": generation.get("strategy", "")},
        {"Item": "Allowed outputs", "Value": ", ".join(generation.get("allowed_outputs", []))},
        {"Item": "Guardrails", "Value": "; ".join(generation.get("guardrails", []))},
    ]

    sheets = {
        "AI Approach": pd.DataFrame(prediction_rows + generation_rows),
        "Business Targets": pd.DataFrame(
            [{"Target": key, "Value": value} for key, value in targets.items()]
        ),
        "Optimization": pd.DataFrame(
            [{"Metric": key, "Value": value} for key, value in optimization.items()]
        ),
        "Top Channels": pd.DataFrame(evidence_packet.get("top_channel_evidence", [])),
        "Recommended Allocation": pd.DataFrame(evidence_packet.get("recommended_allocation", [])),
    }
    evaluation = evidence_packet.get("evaluation", {}) or {}
    if evaluation:
        sheets["Evaluation"] = pd.DataFrame([{"Metric": key, "Value": value} for key, value in evaluation.items()])
    return sheets


def build_evidence_workbook(evidence_packet: Mapping[str, object]) -> bytes:
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for sheet_name, frame in _evidence_packet_sheets(evidence_packet).items():
            _write_excel_frame(writer, frame, sheet_name)
        _style_workbook(writer.book)
    return buffer.getvalue()


def build_allocation_workbook(
    optimization: Mapping[str, object],
    scorecard: pd.DataFrame,
    recommendations: list[str],
    pilot_plan: pd.DataFrame | None = None,
    risk_audit: pd.DataFrame | None = None,
) -> bytes:
    summary = pd.DataFrame(
        [
            {"Metric": "Current budget", "Value": optimization.get("current_budget")},
            {"Metric": "Recommended budget", "Value": optimization.get("recommended_budget")},
            {"Metric": "Optimized revenue", "Value": optimization.get("optimized_revenue")},
            {"Metric": "Revenue delta", "Value": optimization.get("revenue_delta")},
            {"Metric": "Revenue delta %", "Value": optimization.get("revenue_delta_pct")},
            {"Metric": "Unallocated budget", "Value": optimization.get("unallocated_budget")},
        ]
    )
    recommendations_df = pd.DataFrame({"Recommendation": recommendations})

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        _write_excel_frame(writer, summary, "Budget Summary")
        _write_excel_frame(writer, display_business_scorecard(scorecard), "KPI Scorecard")
        _write_excel_frame(writer, optimization["allocation"], "Recommended Allocation")
        _write_excel_frame(writer, recommendations_df, "Recommendations")
        if pilot_plan is not None:
            _write_excel_frame(writer, pilot_plan, "Pilot Plan")
        if risk_audit is not None:
            _write_excel_frame(writer, risk_audit, "Responsible AI Audit")
        _style_workbook(writer.book)
    return buffer.getvalue()


def build_executive_report_pdf(
    data: pd.DataFrame,
    contribution: pd.DataFrame,
    optimization: Mapping[str, object],
    simulation: Mapping[str, object],
    evaluation: Mapping[str, object] | None,
    recommendations: list[str],
    scorecard: pd.DataFrame,
    max_mape_pct: float = 15.0,
) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buffer = io.BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=0.55 * inch,
        leftMargin=0.55 * inch,
        topMargin=0.55 * inch,
        bottomMargin=0.55 * inch,
        title="Mixalyzer Executive Report",
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "MixalyzerTitle",
        parent=styles["Title"],
        textColor=colors.HexColor("#0F766E"),
        fontSize=22,
        leading=26,
        spaceAfter=8,
    )
    section_style = ParagraphStyle(
        "MixalyzerSection",
        parent=styles["Heading2"],
        textColor=colors.HexColor("#0F172A"),
        fontSize=13,
        leading=16,
        spaceBefore=10,
        spaceAfter=6,
    )
    body_style = ParagraphStyle(
        "MixalyzerBody",
        parent=styles["BodyText"],
        fontSize=9,
        leading=12,
        textColor=colors.HexColor("#1F2937"),
    )

    def table(data_rows: list[list[object]], widths: list[float] | None = None) -> Table:
        formatted_rows = [
            [Paragraph(escape(str(cell)), body_style) for cell in row]
            for row in data_rows
        ]
        built = Table(formatted_rows, colWidths=widths, hAlign="LEFT")
        built.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F766E")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#CBD5E1")),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F8FAFC")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ]
            )
        )
        return built

    total_spend_value = float(data[list(DEFAULT_CHANNELS)].sum().sum())
    total_revenue_value = float(data[TARGET_COL].sum())
    total_customers_value = float(data[CUSTOMER_COL].sum()) if CUSTOMER_COL in data.columns else 0.0
    cac_value = total_spend_value / total_customers_value if total_customers_value else None
    top_roi = contribution.sort_values("ROI", ascending=False).iloc[0]
    weakest_roi = contribution.sort_values("ROI", ascending=True).iloc[0]
    allocation = optimization["allocation"][
        ["Channel", "Current Spend", "Recommended Spend", "Spend Shift", "Change %"]
    ].copy()
    trust = model_trust_assessment(evaluation, max_mape_pct)
    summary = executive_recommendation_summary(optimization, scorecard, evaluation, max_mape_pct)
    risk_audit = build_responsible_ai_audit(data, evaluation, contribution, max_mape_pct)
    pilot_plan = build_pilot_plan(optimization, scorecard, evaluation, max_mape_pct)

    story = [
        Paragraph("Mixalyzer Executive Report", title_style),
        Paragraph(f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}", body_style),
        Spacer(1, 8),
        Paragraph("Business Context", section_style),
        table(
            [
                ["Topic", "Detail"],
                ["Business problem", "Growth teams waste marketing budget when they cannot identify which channels drive incremental revenue."],
                ["Target customer", "Growth teams, marketing analysts, finance teams, CMOs, and D2C/e-commerce companies."],
                ["Value proposition", "Upload weekly data, evaluate model quality, simulate budget changes, optimize allocation, and export recommendations."],
            ],
            [1.6 * inch, 4.8 * inch],
        ),
        Paragraph("Who / Why / How", section_style),
        table(
            [
                ["Framework", "Mixalyzer answer"],
                ["Who", "Marketing analysts, growth managers, finance teams, CMOs, and executive stakeholders."],
                ["Why", "Last-click tools miss delayed effects, seasonality, offline channels, diminishing returns, and cross-channel impact."],
                ["How", "MMM with adstock, saturation, seasonality, optional event controls, baseline comparison, simulation, optimization, and executive exports."],
            ],
            [1.2 * inch, 5.2 * inch],
        ),
        Paragraph("Business KPI Snapshot", section_style),
        table(
            [
                ["Metric", "Value"],
                ["Revenue analyzed", money(total_revenue_value)],
                ["Marketing spend analyzed", money(total_spend_value)],
                ["Marketing ROI", f"{total_revenue_value / total_spend_value:.2f}x" if total_spend_value else "N/A"],
                ["CAC", money(cac_value) if cac_value else "N/A"],
                ["Predicted optimized lift", f"{signed_money(optimization['revenue_delta'])} ({optimization['revenue_delta_pct']:.1f}%)"],
                ["Confidence range", f"{signed_money(optimization['revenue_delta_low'])} to {signed_money(optimization['revenue_delta_high'])}"],
                ["Model trust badge", trust["label"]],
            ],
            [2.3 * inch, 4.1 * inch],
        ),
        Paragraph("Executive Recommendation Summary", section_style),
        table(
            [
                ["Field", "Summary"],
                ["Recommendation", summary["recommendation"]],
                ["Expected impact", summary["impact"]],
                ["Confidence", summary["confidence"]],
                ["Risk", summary["risk"]],
                ["Next step", summary["next_step"]],
            ],
            [1.45 * inch, 4.95 * inch],
        ),
        Paragraph("Business Goal Attainment", section_style),
        table(
            [["Business KPI", "Current", "Projected", "Delta", "Target", "Status"]]
            + display_business_scorecard(scorecard)[
                ["Business KPI", "Current", "Projected", "Delta", "Target", "Status"]
            ].values.tolist()
        ),
        Paragraph("Recommended Allocation", section_style),
        table(
            [["Channel", "Current Spend", "Recommended Spend", "Shift", "Change"]]
            + [
                [
                    row["Channel"],
                    money(float(row["Current Spend"])),
                    money(float(row["Recommended Spend"])),
                    signed_money(float(row["Spend Shift"])),
                    pct(float(row["Change %"])),
                ]
                for _, row in allocation.iterrows()
            ]
        ),
        Paragraph("Channel Insights", section_style),
        table(
            [
                ["Insight", "Value"],
                ["Highest estimated ROI", f"{top_roi['Channel']} at {float(top_roi['ROI']):.2f}x"],
                ["Lowest estimated ROI", f"{weakest_roi['Channel']} at {float(weakest_roi['ROI']):.2f}x"],
                ["Active simulation impact", f"{signed_money(simulation['revenue_delta'])} ({simulation['revenue_delta_pct']:.1f}%)"],
            ],
            [2.3 * inch, 4.1 * inch],
        ),
    ]

    story.append(Paragraph("Model Evaluation", section_style))
    if evaluation:
        story.append(
            table(
                [
                    ["Metric", "Value"],
                    ["Train rows", evaluation["train_rows"]],
                    ["Test rows", evaluation["test_rows"]],
                    ["MMX MAPE", f"{evaluation['model_metrics']['mape']:.2f}%"],
                    ["Baseline MAPE", f"{evaluation['baseline_metrics']['mape']:.2f}%"],
                    ["RMSE improvement", f"{evaluation['rmse_improvement_pct']:.1f}%"],
                ],
                [2.3 * inch, 4.1 * inch],
            )
        )
    else:
        story.append(Paragraph("Evaluation unavailable because the uploaded dataset is too short.", body_style))

    story.append(Paragraph("Management Recommendations", section_style))
    for recommendation in recommendations:
        story.append(Paragraph(f"- {escape(recommendation)}", body_style))

    story.append(Paragraph("Responsible AI Notes", section_style))
    story.append(
        table(
            [["Risk Area", "Example", "Impact", "Mitigation", "Status"]]
            + risk_audit[
                ["Risk Area", "Example in Mixalyzer", "Potential Impact", "Mitigation / Guardrail", "Status"]
            ].values.tolist(),
            [0.8 * inch, 1.45 * inch, 1.35 * inch, 1.95 * inch, 0.85 * inch],
        )
    )

    story.append(Paragraph("Rollout / Pilot Plan", section_style))
    story.append(
        table(
            [["Plan Area", "Action"]] + pilot_plan.values.tolist(),
            [1.6 * inch, 4.8 * inch],
        )
    )

    story.append(Paragraph("Caveats And Next Steps", section_style))
    for note in [
        "Treat recommendations as decision support, not automated media-buying instructions.",
        "Validate large reallocations with incrementality tests or controlled experiments.",
        "Use aggregated spend, revenue, and customer data; avoid customer-level identifiers.",
        "Refresh the model as new weekly data arrives and compare results with actual lift.",
    ]:
        story.append(Paragraph(f"- {note}", body_style))

    document.build(story)
    return buffer.getvalue()


def line_spend_revenue_chart(data: pd.DataFrame) -> alt.Chart:
    frame = data.copy()
    frame["total_spend"] = frame[list(DEFAULT_CHANNELS)].sum(axis=1)
    long = frame.melt(
        id_vars=["date"],
        value_vars=[TARGET_COL, "total_spend"],
        var_name="Metric",
        value_name="Value",
    )
    long["Metric"] = long["Metric"].map({"revenue": "Revenue", "total_spend": "Marketing Spend"})

    return (
        alt.Chart(long)
        .mark_line(strokeWidth=3)
        .encode(
            x=alt.X("date:T", title=None),
            y=alt.Y("Value:Q", title=None),
            color=alt.Color(
                "Metric:N",
                scale=alt.Scale(range=["#38d7c1", "#7c8cff"]),
                legend=alt.Legend(orient="top"),
            ),
            tooltip=[
                alt.Tooltip("date:T", title="Week"),
                alt.Tooltip("Metric:N"),
                alt.Tooltip("Value:Q", format="$,.0f"),
            ],
        )
        .properties(height=320)
    )


def channel_spend_chart(data: pd.DataFrame) -> alt.Chart:
    long = data.melt(
        id_vars=["date"],
        value_vars=list(DEFAULT_CHANNELS),
        var_name="channel",
        value_name="Spend",
    )
    long["Channel"] = long["channel"].map(CHANNEL_LABELS)
    return (
        alt.Chart(long)
        .mark_area(opacity=0.88)
        .encode(
            x=alt.X("date:T", title=None),
            y=alt.Y("Spend:Q", stack=True, title="Weekly spend"),
            color=alt.Color("Channel:N", legend=alt.Legend(orient="bottom")),
            tooltip=[
                alt.Tooltip("date:T", title="Week"),
                alt.Tooltip("Channel:N"),
                alt.Tooltip("Spend:Q", format="$,.0f"),
            ],
        )
        .properties(height=280)
    )


def roi_chart(contribution: pd.DataFrame) -> alt.Chart:
    return (
        alt.Chart(contribution)
        .mark_bar(cornerRadiusTopRight=4, cornerRadiusBottomRight=4)
        .encode(
            y=alt.Y("Channel:N", sort="-x", title=None),
            x=alt.X("ROI:Q", title="Estimated revenue per spend dollar"),
            color=alt.Color(
                "ROI:Q",
                scale=alt.Scale(range=["#f97316", "#38d7c1"]),
                legend=None,
            ),
            tooltip=[
                alt.Tooltip("Channel:N"),
                alt.Tooltip("Spend:Q", format="$,.0f"),
                alt.Tooltip("Estimated Contribution:Q", format="$,.0f"),
                alt.Tooltip("ROI:Q", format=".2f"),
            ],
        )
        .properties(height=300)
    )


def contribution_chart(contribution: pd.DataFrame) -> alt.Chart:
    return (
        alt.Chart(contribution)
        .mark_arc(innerRadius=62, outerRadius=122)
        .encode(
            theta=alt.Theta("Estimated Contribution:Q"),
            color=alt.Color("Channel:N", legend=alt.Legend(orient="bottom")),
            tooltip=[
                alt.Tooltip("Channel:N"),
                alt.Tooltip("Estimated Contribution:Q", format="$,.0f"),
                alt.Tooltip("Contribution Share:Q", format=".1%"),
            ],
        )
        .properties(height=300)
    )


def spend_shift_chart(details: pd.DataFrame, current_col: str, scenario_col: str) -> alt.Chart:
    long = details.melt(
        id_vars=["Channel"],
        value_vars=[current_col, scenario_col],
        var_name="Scenario",
        value_name="Spend",
    )
    return (
        alt.Chart(long)
        .mark_bar(cornerRadiusTopLeft=4, cornerRadiusTopRight=4)
        .encode(
            x=alt.X("Channel:N", title=None),
            y=alt.Y("Spend:Q", title="Weekly spend"),
            color=alt.Color(
                "Scenario:N",
                scale=alt.Scale(range=["#94a3b8", "#38d7c1"]),
                legend=alt.Legend(orient="top"),
            ),
            tooltip=[
                alt.Tooltip("Channel:N"),
                alt.Tooltip("Scenario:N"),
                alt.Tooltip("Spend:Q", format="$,.0f"),
            ],
        )
        .properties(height=310)
    )


def response_curve_chart(curve: pd.DataFrame) -> alt.Chart:
    return (
        alt.Chart(curve)
        .mark_line(point=True, strokeWidth=3)
        .encode(
            x=alt.X("Spend:Q", title="Weekly spend"),
            y=alt.Y("Predicted Revenue:Q", title="Predicted revenue"),
            tooltip=[
                alt.Tooltip("Spend:Q", format="$,.0f"),
                alt.Tooltip("Predicted Revenue:Q", format="$,.0f"),
                alt.Tooltip("Incremental Revenue:Q", format="$,.0f"),
            ],
        )
        .properties(height=310)
    )


def build_grounded_ai_payload(
    contribution: pd.DataFrame,
    optimization: Mapping[str, object],
    simulation: Mapping[str, object],
    evidence_packet: Mapping[str, object],
    scorecard: pd.DataFrame,
    trust: Mapping[str, object],
    readiness: Mapping[str, object],
    risk_audit: pd.DataFrame,
    recommendation_summary: Mapping[str, str],
) -> dict[str, object]:
    allocation = optimization["allocation"].copy()
    readiness_checks = readiness.get("checks", pd.DataFrame())
    readiness_warnings = []
    if isinstance(readiness_checks, pd.DataFrame) and not readiness_checks.empty:
        readiness_warnings = (
            readiness_checks.loc[readiness_checks["Status"] != "Ready"]
            .head(5)
            .to_dict("records")
        )

    shifts = allocation_shift_summary(optimization)
    increase_rows = shifts["increase_rows"]
    decrease_rows = shifts["decrease_rows"]
    return {
        "workflow": [
            "data upload",
            "readiness checks",
            "MMM model training",
            "baseline comparison",
            "channel contribution and ROI analysis",
            "simulation",
            "optimization",
            "evidence packet",
            "AI-generated stakeholder recommendation",
            "responsible AI/risk review",
            "rollout plan",
        ],
        "executive_summary": dict(recommendation_summary),
        "model_trust": dict(trust),
        "model_metrics": evidence_packet.get("prediction_layer", {}).get("metrics", {}),
        "evaluation": evidence_packet.get("evaluation", {}),
        "business_targets": evidence_packet.get("business_targets", {}),
        "kpi_scorecard": display_business_scorecard(scorecard).to_dict("records"),
        "top_channel_evidence": evidence_packet.get("top_channel_evidence", []),
        "recommended_allocation": allocation[
            ["Channel", "Current Spend", "Recommended Spend", "Spend Shift", "Change %"]
        ].to_dict("records"),
        "channels_to_increase": (
            increase_rows[["Channel", "Spend Shift", "Change %"]].head(3).to_dict("records")
            if isinstance(increase_rows, pd.DataFrame) and not increase_rows.empty
            else []
        ),
        "channels_to_decrease": (
            decrease_rows[["Channel", "Spend Shift", "Change %"]].head(3).to_dict("records")
            if isinstance(decrease_rows, pd.DataFrame) and not decrease_rows.empty
            else []
        ),
        "optimization": evidence_packet.get("optimization", {}),
        "confidence_range": {
            "low": round(float(optimization.get("revenue_delta_low", 0.0) or 0.0), 2),
            "expected_pct": round(float(optimization.get("revenue_delta_pct", 0.0) or 0.0), 2),
            "high": round(float(optimization.get("revenue_delta_high", 0.0) or 0.0), 2),
        },
        "active_simulation": {
            "revenue_delta_pct": round(float(simulation.get("revenue_delta_pct", 0.0) or 0.0), 2),
            "budget_delta": round(float(simulation.get("budget_delta", 0.0) or 0.0), 2),
        },
        "readiness": {
            "score": readiness.get("score"),
            "status": readiness.get("status"),
            "warnings": readiness_warnings,
        },
        "risk_audit": risk_audit.to_dict("records"),
        "guardrails": [
            "Use only model evidence and KPI results in the recommendation.",
            "Mention model trust, MAPE, confidence range, and caveats.",
            "Recommend a pilot when model quality or conservative impact is uncertain.",
            "Require human review before permanent budget changes.",
        ],
    }


def grounded_ai_brief_html(payload: Mapping[str, object]) -> str:
    summary = payload.get("executive_summary", {}) or {}
    trust = payload.get("model_trust", {}) or {}
    confidence = payload.get("confidence_range", {}) or {}
    readiness = payload.get("readiness", {}) or {}
    low_value = confidence.get("low", 0.0)
    high_value = confidence.get("high", 0.0)
    try:
        confidence_range = f"{signed_money(float(low_value))} to {signed_money(float(high_value))}"
    except (TypeError, ValueError):
        confidence_range = "N/A"
    return f"""
    <div class="ai-brief">
      <h4>Grounded AI stakeholder brief</h4>
      <p><strong>Recommended action:</strong> {escape(str(summary.get("recommendation", "")))}</p>
      <p><strong>Expected business impact:</strong> {escape(str(summary.get("impact", "")))}</p>
      <p><strong>Model confidence:</strong> {escape(str(summary.get("confidence", "")))} Trust status is {escape(str(trust.get("label", "Unknown")))}.</p>
      <p><strong>Evidence used:</strong> MMM metrics, baseline comparison, channel ROI/contribution, optimization output, KPI scorecard, and responsible AI audit.</p>
      <p><strong>Data and risk caveat:</strong> Readiness is {escape(str(readiness.get("score", "N/A")))}/100 ({escape(str(readiness.get("status", "Unknown")))}); conservative-to-optimistic revenue range is {escape(confidence_range)}.</p>
      <p><strong>Next step:</strong> {escape(str(summary.get("next_step", "")))}</p>
    </div>
    """


@st.cache_data(show_spinner=False, ttl=900)
def maybe_generate_openai_recommendations(
    payload: Mapping[str, object],
) -> str | None:
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        return None

    try:
        from openai import OpenAI

        client = OpenAI(api_key=api_key)
        response = client.chat.completions.create(
            model=os.getenv("OPENAI_MMX_MODEL", os.getenv("OPENAI_MODEL", "gpt-4o-mini")),
            messages=[
                {
                    "role": "system",
                    "content": (
                        "You are Mixalyzer's grounded recommendation engine for marketing mix optimization. "
                        "Use only the JSON evidence payload. Do not invent channels, metrics, dates, or impact numbers. "
                        "Write a concise stakeholder-ready recommendation with these sections: Recommended action, "
                        "Expected business impact, Evidence used, Confidence and caveats, Pilot next step. "
                        "Mention model trust, MAPE or baseline comparison when available, confidence range, and human review. "
                        "If model trust is weak, lead with caution and recommend improving data/model quality before acting."
                    ),
                },
                {"role": "user", "content": json.dumps(payload, default=str)},
            ],
            temperature=0.2,
        )
        return response.choices[0].message.content or None
    except Exception:
        return None


with st.sidebar:
    sidebar_logo_uri = asset_data_uri("mixalyzer_logo.svg")
    if sidebar_logo_uri:
        st.markdown(
            f"<img src='{sidebar_logo_uri}' alt='Mixalyzer logo' style='width:100%;max-width:245px;margin-bottom:0.75rem;'>",
            unsafe_allow_html=True,
        )
    else:
        st.markdown("### Mixalyzer")
    uploaded = st.file_uploader("Marketing dataset", type=["csv"])
    st.download_button(
        "Download CSV template",
        data=build_csv_template(),
        file_name="marketing_mix_template.csv",
        mime="text/csv",
        key="download_csv_template",
        width="stretch",
    )
    raw_data = read_input_data(uploaded)
    with st.expander("Auto column mapping", expanded=uploaded is not None):
        column_mapping = build_column_mapping_controls(raw_data, uploaded is not None)
    mapped_input = apply_column_mapping(raw_data, column_mapping) if uploaded is not None else raw_data
    readiness = assess_data_readiness(mapped_input)
    st.caption(f"Data readiness: {readiness['score']}/100 · {readiness['status']}")
    model_engine = st.selectbox(
        "Model engine",
        ["Ridge MMM (fast)", "Bayesian MMM (posterior)"],
        index=0,
    )
    confidence_level = st.select_slider(
        "Confidence range",
        options=[0.80, 0.90, 0.95],
        value=0.80,
        format_func=lambda value: f"{int(value * 100)}%",
    )
    regularization = st.slider("Model regularization", 0.1, 5.0, 1.5, 0.1)
    use_openai = bool(os.getenv("OPENAI_API_KEY"))
    st.caption("AI recommendations are generated automatically from the MMM evidence workflow.")
    with st.expander("Business KPI targets", expanded=True):
        target_roi_lift_pct = st.slider("Target ROI lift", 0.0, 25.0, 5.0, 0.5, format="%.1f%%")
        target_cac_reduction_pct = st.slider("Target CAC reduction", 0.0, 25.0, 3.0, 0.5, format="%.1f%%")
        max_mape_pct = st.slider("Maximum planning MAPE", 1.0, 40.0, 15.0, 0.5, format="%.1f%%")


try:
    with st.spinner("Mixalyzer is cleaning data and training the MMM engine..."):
        data = prepare_marketing_data(mapped_input)
        model, contribution_df, baseline_scenario = train_model(data, regularization, model_engine)
except Exception as exc:
    st.error(f"Could not train the marketing mix model: {exc}")
    st.stop()

try:
    with st.spinner("Running holdout evaluation and model comparison..."):
        evaluation_results = evaluate_current_model(data, regularization)
        model_comparison_df = compare_models(data, regularization)
except Exception:
    evaluation_results = None
    model_comparison_df = pd.DataFrame()


total_spend = float(data[list(DEFAULT_CHANNELS)].sum().sum())
total_revenue = float(data[TARGET_COL].sum())
model_roi = float(total_revenue / total_spend) if total_spend else 0.0
total_customers = float(data[CUSTOMER_COL].sum()) if CUSTOMER_COL in data.columns else 0.0
cac = float(total_spend / total_customers) if total_customers else None
current_weekly_budget = sum(float(baseline_scenario[channel]) for channel in DEFAULT_CHANNELS)
default_optimization = optimize_with_confidence(
    model,
    baseline_scenario,
    float(current_weekly_budget),
    confidence_level,
)
target_summary = {
    "target_roi_lift_pct": float(target_roi_lift_pct),
    "target_cac_reduction_pct": float(target_cac_reduction_pct),
    "max_mape_pct": float(max_mape_pct),
}
business_scorecard = build_business_kpi_scorecard(
    data,
    default_optimization,
    model.metrics,
    target_roi_lift_pct=target_roi_lift_pct,
    target_cac_reduction_pct=target_cac_reduction_pct,
    max_mape_pct=max_mape_pct,
)
evidence_packet = build_genai_evidence_packet(
    model,
    contribution_df,
    default_optimization,
    evaluation_results,
    target_summary,
)
trust_assessment = model_trust_assessment(evaluation_results, max_mape_pct)
default_summary = executive_recommendation_summary(
    default_optimization,
    business_scorecard,
    evaluation_results,
    max_mape_pct,
)

logo_uri = asset_data_uri("mixalyzer_logo.svg")
hero_uri = asset_data_uri("mixalyzer_hero.svg")
logo_markup = (
    f"<img class='brand-logo' src='{logo_uri}' alt='Mixalyzer logo'>"
    if logo_uri
    else "<div class='brand-mark'>MX</div>"
)

st.markdown(
    f"""
    <div class="app-header">
      <div class="brand-lockup">
        {logo_markup}
        <div class="brand-copy">
          <div class="brand-kicker">Marketing intelligence platform</div>
          <div class="brand-title">AI-powered marketing mix optimization for growth teams</div>
          <div class="brand-subtitle">
            Upload weekly marketing data, evaluate model quality, simulate budget changes, optimize allocation,
            and export stakeholder-ready recommendations tied to ROI, CAC, revenue lift, MAPE, and confidence ranges.
          </div>
        </div>
      </div>
      <div class="brand-pill">MMM + Optimization + Grounded GenAI</div>
    </div>
    """,
    unsafe_allow_html=True,
)

kpi_cols = st.columns(6)
top_metrics = [
    ("Revenue", money(total_revenue)),
    ("Marketing spend", money(total_spend)),
    ("Marketing ROI", f"{model_roi:.2f}x"),
    ("CAC", money(cac) if cac is not None else "N/A"),
    ("Model R-squared", f"{model.metrics['r2']:.2f}"),
    ("MAPE", pct(model.metrics["mape"])),
]
for col, (label, value) in zip(kpi_cols, top_metrics):
    with col:
        render_metric_card(label, value)

(
    home_tab,
    strategy_tab,
    business_goals_tab,
    data_setup_tab,
    dashboard_tab,
    simulation_tab,
    optimization_tab,
    evaluation_tab,
    responsible_ai_tab,
    pilot_plan_tab,
    model_tab,
) = st.tabs(
    [
        "Home",
        "Strategy",
        "Business Goals",
        "Data Setup",
        "Dashboard",
        "Simulation",
        "Optimization",
        "Evaluation",
        "Responsible AI",
        "Pilot Plan",
        "Model",
    ]
)

with home_tab:
    hero_visual_markup = (
        f"<img class='hero-visual' src='{hero_uri}' alt='Mixalyzer dashboard and AI recommendation illustration'>"
        if hero_uri
        else """
        <div class="feature-card">
          <h4>Hero image placeholder</h4>
          <p>Add a replacement image at assets/mixalyzer_hero.svg or update the asset path in mmx_app.py.</p>
        </div>
        """
    )
    st.markdown(
        f"""
        <div class="hero">
          <div>
            <div class="hero-eyebrow">Marketing intelligence. Budget optimization. AI decision support.</div>
            <h2>Turn messy channel spend into a confident budget plan.</h2>
            <p>
              Mixalyzer combines MMM prediction, baseline comparison, optimization, and grounded AI narrative generation
              so growth, finance, and executive teams can see what changed, what to do next, and how much risk remains.
            </p>
            <div class="identity-row">
              <span class="identity-chip">Channel contribution</span>
              <span class="identity-chip">Marginal response</span>
              <span class="identity-chip">Executive recommendations</span>
            </div>
          </div>
          <div>{hero_visual_markup}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    render_executive_summary(default_summary)

    home_cols = st.columns(3)
    with home_cols[0]:
        st.markdown(
            """
            <div class="feature-card">
              <h4>Target customer</h4>
              <p>Growth teams, marketing analysts, finance teams, CMOs, and D2C/e-commerce companies.</p>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with home_cols[1]:
        st.markdown(
            """
            <div class="feature-card">
              <h4>Business problem</h4>
              <p>Teams waste budget when they cannot identify which channels drive incremental revenue.</p>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with home_cols[2]:
        st.markdown(
            """
            <div class="feature-card">
              <h4>Value proposition</h4>
              <p>Upload weekly data, evaluate model quality, simulate spend, optimize allocation, and export recommendations.</p>
            </div>
            """,
            unsafe_allow_html=True,
        )

    st.divider()
    product_cols = st.columns(5)
    product_metrics = [
        ("ROI lift", f"{target_roi_lift_pct:.1f}% target"),
        ("CAC reduction", f"{target_cac_reduction_pct:.1f}% target"),
        ("Revenue lift", pct(default_optimization["revenue_delta_pct"])),
        ("Forecast error", f"MAPE <= {pct(max_mape_pct)}"),
        ("Confidence range", f"{int(confidence_level * 100)}%"),
    ]
    for col, (label, value) in zip(product_cols, product_metrics):
        with col:
            render_metric_card(label, value)

    cta_cols = st.columns(3)
    with cta_cols[0]:
        st.button("Review data readiness", width="stretch")
    with cta_cols[1]:
        st.button("Run budget optimization", width="stretch")
    with cta_cols[2]:
        st.button("Export executive outputs", width="stretch")

    feature_cols = st.columns(4)
    feature_copy = [
        ("Data readiness", "Check whether the uploaded dataset is clean enough for MMX."),
        ("Model trust", "Compare MMM against baselines before relying on recommendations."),
        ("Optimization", "Turn model outputs into a next-period budget plan."),
        ("Executive outputs", "Download a PDF report, allocation workbook, and evidence workbook."),
    ]
    for col, (title, body) in zip(feature_cols, feature_copy):
        with col:
            st.markdown(
                f"""
                <div class="feature-card">
                  <h4>{title}</h4>
                  <p>{body}</p>
                </div>
                """,
                unsafe_allow_html=True,
            )

    st.subheader("Why AI is appropriate here")
    ai_cols = st.columns(4)
    ai_cards = [
        (
            "MMM estimates impact",
            "The ML layer learns channel contribution and marginal response from historical weekly spend and revenue.",
        ),
        (
            "Optimization chooses a plan",
            "The optimizer reallocates budget under constraints instead of only ranking channels after the fact.",
        ),
        (
            "GenAI translates evidence",
            "The narrative layer uses model metrics, KPI results, confidence ranges, and the evidence packet to write for stakeholders.",
        ),
        (
            "Humans stay in control",
            "Responsible AI checks and pilot rules keep recommendations as decision support before budget changes go live.",
        ),
    ]
    for col, (title, body) in zip(ai_cols, ai_cards):
        with col:
            st.markdown(
                f"""
                <div class="feature-card">
                  <h4>{title}</h4>
                  <p>{body}</p>
                </div>
                """,
                unsafe_allow_html=True,
            )

with strategy_tab:
    st.subheader("Business case and strategy")
    who_col, why_col, how_col = st.columns(3)
    with who_col:
        st.markdown(
            """
            <div class="feature-card">
              <h4>Who</h4>
              <p><strong>Primary users:</strong> marketing analysts, growth managers, finance teams, and CMOs.</p>
              <p><strong>Target companies:</strong> small-to-mid-size e-commerce, D2C, subscription, and retail businesses.</p>
              <p><strong>Stakeholders:</strong> marketing, finance, sales, and executive leadership.</p>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with why_col:
        st.markdown(
            """
            <div class="feature-card">
              <h4>Why</h4>
              <p>Last-click tools over-focus on immediate digital clicks and miss delayed effects, seasonality, offline channels, diminishing returns, and cross-channel impact.</p>
              <p>Marketing leaders need budget recommendations tied to revenue, ROI, CAC, and risk.</p>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with how_col:
        st.markdown(
            """
            <div class="feature-card">
              <h4>How</h4>
              <p>Mixalyzer uses weekly spend, revenue, customers/conversions, adstock, saturation, seasonality, optional event controls, and baseline comparison.</p>
              <p>It supports simulation, optimization, responsible AI checks, executive exports, and business recommendation generation.</p>
            </div>
            """,
            unsafe_allow_html=True,
        )

    st.subheader("Integrated AI decision workflow")
    st.markdown(
        """
        <div class="summary-box">
          <h4>From data to decision support</h4>
          <p>Data upload -> readiness checks -> MMM model training -> baseline comparison -> channel contribution and ROI analysis -> simulation -> optimization -> evidence packet -> AI-generated stakeholder recommendation -> responsible AI/risk review -> rollout plan.</p>
          <p>Generative AI is the translation layer, not the modeling layer: it reads the MMM evidence, KPI impact, confidence range, and risk audit before writing a business recommendation.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.subheader("Competitive advantage")
    advantage_cols = st.columns(5)
    advantages = [
        ("Better budget allocation", "Reallocate spend toward channels with stronger marginal response."),
        ("Faster planning cycles", "Move from CSV upload to budget recommendation in one workflow."),
        ("Transparent decisions", "Show model quality, confidence ranges, and risk caveats."),
        ("Marketing-finance alignment", "Tie recommendations to ROI, CAC, revenue, and MAPE."),
        ("Continuous monitoring", "Use weekly cadence instead of one-time analysis."),
    ]
    for col, (title, body) in zip(advantage_cols, advantages):
        with col:
            st.markdown(
                f"""
                <div class="feature-card">
                  <h4>{title}</h4>
                  <p>{body}</p>
                </div>
                """,
                unsafe_allow_html=True,
            )

with business_goals_tab:
    st.subheader("Business KPI targets")
    goal_cols = st.columns(4)
    goal_metrics = [
        ("Primary KPI", "Marketing ROI", "Revenue per spend dollar"),
        ("ROI lift target", pct(target_roi_lift_pct), "Optimized vs current mix"),
        ("CAC reduction target", pct(target_cac_reduction_pct), "Requires customer data"),
        ("Planning error target", f"MAPE <= {pct(max_mape_pct)}", "Train/test forecast quality"),
    ]
    for col, (label, value, delta) in zip(goal_cols, goal_metrics):
        with col:
            render_metric_card(label, value, delta)

    st.subheader("Goal attainment from the current optimized plan")
    scorecard_display = display_business_scorecard(business_scorecard)
    status_cols = st.columns(3)
    for col, metric_name in zip(status_cols, ["Marketing ROI lift", "CAC reduction", "Forecast quality"]):
        row = scorecard_display.loc[scorecard_display["Business KPI"] == metric_name].iloc[0]
        with col:
            render_metric_card(metric_name, row["Status"], row["Delta"])

    st.dataframe(scorecard_display, hide_index=True, width="stretch")

    st.subheader("Selected AI approach")
    approach_cols = st.columns(3)
    approach_cards = [
        (
            "Prediction layer",
            f"{model.model_kind} predicts revenue from spend, seasonality, adstock carryover, and diminishing returns.",
        ),
        (
            "Generative layer",
            "Grounded recommendation language is conditioned on the MMM evidence packet, not free-form guesses.",
        ),
        (
            "Business decision layer",
            "Budget changes are measured against ROI lift, CAC reduction, forecast error, and human review.",
        ),
    ]
    for col, (title, body) in zip(approach_cols, approach_cards):
        with col:
            st.markdown(
                f"""
                <div class="feature-card">
                  <h4>{title}</h4>
                  <p>{body}</p>
                </div>
                """,
                unsafe_allow_html=True,
            )

    st.download_button(
        "Download evidence workbook",
        data=build_evidence_workbook(evidence_packet),
        file_name="mixalyzer_genai_evidence_packet.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="download_business_goals_evidence_packet",
        width="stretch",
    )

with data_setup_tab:
    st.subheader("Data readiness and column mapping")
    ready_cols = st.columns(4)
    with ready_cols[0]:
        render_metric_card("Readiness score", f"{readiness['score']}/100")
    with ready_cols[1]:
        render_metric_card("Status", readiness["status"])
    with ready_cols[2]:
        render_metric_card("Rows", f"{len(data):,}")
    with ready_cols[3]:
        render_metric_card("Mapped fields", f"{len([v for v in column_mapping.values() if v])}/{len(column_mapping)}")

    st.markdown(
        f"""
        <div class="feature-card {readiness_class(readiness['status'])}">
          <h4>Data readiness verdict</h4>
          <p>Mixalyzer checked required fields, date quality, history length, numeric quality, CAC support, and spend coverage.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    map_left, map_right = st.columns([1, 1])
    with map_left:
        mapping_rows = [
            {
                "Mixalyzer Field": FIELD_LABELS.get(canonical, canonical),
                "Source Column": source or "Not mapped",
            }
            for canonical, source in column_mapping.items()
        ]
        st.dataframe(pd.DataFrame(mapping_rows), hide_index=True, width="stretch")
    with map_right:
        st.dataframe(readiness["checks"], hide_index=True, width="stretch")

    st.subheader("External event controls")
    st.caption(
        "Controls help prevent marketing channels from receiving credit or blame for events such as holidays, stockouts, launches, competitor campaigns, or macro conditions."
    )
    st.dataframe(detected_control_summary(data), hide_index=True, width="stretch")

    st.subheader("Cleaned data preview")
    st.dataframe(data.head(20), hide_index=True, width="stretch")

with dashboard_tab:
    st.subheader("What happened historically?")
    st.caption("Use this view to compare revenue movement against total marketing spend over time.")
    left, right = st.columns([1.45, 1])
    with left:
        st.subheader("Spend and revenue trend")
        st.altair_chart(line_spend_revenue_chart(data), use_container_width=True)
        st.caption("Business interpretation: look for periods where revenue moved with spend, and periods where external factors or diminishing returns may explain gaps.")
    with right:
        st.subheader("Channel contribution")
        st.altair_chart(contribution_chart(contribution_df), use_container_width=True)
        st.caption("Business interpretation: larger slices indicate channels with higher estimated revenue contribution, not necessarily highest efficiency.")

    st.subheader("Which channels appear most efficient?")
    chart_a, chart_b = st.columns([1, 1])
    with chart_a:
        st.subheader("ROI by channel")
        st.altair_chart(roi_chart(contribution_df), use_container_width=True)
        top_roi_row = contribution_df.sort_values("ROI", ascending=False).iloc[0]
        st.caption(
            f"Business interpretation: {top_roi_row['Channel']} currently shows the highest estimated ROI, but trust depends on model quality and confidence range."
        )
    with chart_b:
        st.subheader("Channel spend mix")
        st.altair_chart(channel_spend_chart(data), use_container_width=True)
        st.caption("Business interpretation: compare spend share to contribution and ROI to identify potential over- or under-investment.")

    st.subheader("How reliable is the model?")
    render_badge(str(trust_assessment["label"]), str(trust_assessment["status"]))
    st.caption(str(trust_assessment["explanation"]))

with simulation_tab:
    st.subheader("What happens if we change spend?")
    st.caption("Adjust channel budgets to test a what-if scenario before changing the media plan.")
    sliders = {}
    slider_cols = st.columns(3)
    for idx, channel in enumerate(DEFAULT_CHANNELS):
        with slider_cols[idx % 3]:
            sliders[channel] = st.slider(
                CHANNEL_LABELS[channel],
                min_value=-60,
                max_value=80,
                value=0,
                step=5,
                format="%d%%",
            )

    simulation = simulate_with_confidence(model, baseline_scenario, sliders, confidence_level)
    metric_cols = st.columns(4)
    simulation_metrics = [
        ("Current revenue", money(simulation["current_revenue"]), None),
        ("Scenario revenue", money(simulation["scenario_revenue"]), pct(simulation["revenue_delta_pct"])),
        ("Budget change", signed_money(simulation["budget_delta"]), None),
        ("Scenario budget", money(simulation["scenario_budget"]), None),
    ]
    for col, (label, value, delta) in zip(metric_cols, simulation_metrics):
        with col:
            render_metric_card(label, value, delta)

    confidence_summary = pd.DataFrame(
        [
            {"Case": "Conservative", "Revenue Delta": simulation["revenue_delta_low"]},
            {"Case": "Expected", "Revenue Delta": simulation["revenue_delta"]},
            {"Case": "Optimistic", "Revenue Delta": simulation["revenue_delta_high"]},
        ]
    )
    st.altair_chart(confidence_range_chart(confidence_summary), use_container_width=True)
    st.caption("Business interpretation: the expected case is the model's central estimate; conservative and optimistic bars show uncertainty around the spend change.")

    sim_left, sim_right = st.columns([1.1, 1])
    with sim_left:
        st.altair_chart(
            spend_shift_chart(simulation["details"], "Current Spend", "Scenario Spend"),
            use_container_width=True,
        )
    with sim_right:
        selected_channel = st.selectbox(
            "Diminishing returns channel",
            list(DEFAULT_CHANNELS),
            format_func=lambda value: CHANNEL_LABELS[value],
        )
        curve = build_response_curve(model, baseline_scenario, selected_channel)
        st.altair_chart(response_curve_chart(curve), use_container_width=True)
        st.caption("Business interpretation: flatter curves indicate diminishing returns, where extra spend produces smaller incremental revenue gains.")

with optimization_tab:
    st.subheader("What budget should we use next?")
    st.caption("Use this page to convert the MMM outputs into an executive-ready budget recommendation.")
    target_budget = st.slider(
        "Next-period marketing budget",
        min_value=float(current_weekly_budget * 0.5),
        max_value=float(current_weekly_budget * 1.5),
        value=float(current_weekly_budget),
        step=500.0,
        format="$%.0f",
    )
    optimization = (
        default_optimization
        if abs(float(target_budget) - float(current_weekly_budget)) < 1e-6
        else optimize_with_confidence(model, baseline_scenario, target_budget, confidence_level)
    )
    optimization_scorecard = build_business_kpi_scorecard(
        data,
        optimization,
        model.metrics,
        target_roi_lift_pct=target_roi_lift_pct,
        target_cac_reduction_pct=target_cac_reduction_pct,
        max_mape_pct=max_mape_pct,
    )
    optimization_evidence_packet = build_genai_evidence_packet(
        model,
        contribution_df,
        optimization,
        evaluation_results,
        target_summary,
    )
    optimization_summary = executive_recommendation_summary(
        optimization,
        optimization_scorecard,
        evaluation_results,
        max_mape_pct,
    )
    render_executive_summary(optimization_summary)

    opt_cols = st.columns(4)
    optimization_metrics = [
        ("Current budget", money(optimization["current_budget"]), None),
        ("Recommended budget", money(optimization["recommended_budget"]), None),
        ("Optimized revenue", money(optimization["optimized_revenue"]), pct(optimization["revenue_delta_pct"])),
        ("Unallocated", money(optimization["unallocated_budget"]), None),
    ]
    for col, (label, value, delta) in zip(opt_cols, optimization_metrics):
        with col:
            render_metric_card(label, value, delta)

    st.caption(
        "Optimized revenue impact range: "
        f"{signed_money(optimization['revenue_delta_low'])} to {signed_money(optimization['revenue_delta_high'])} "
        f"at {int(confidence_level * 100)}% confidence."
    )

    st.subheader("Business KPI impact")
    st.dataframe(display_business_scorecard(optimization_scorecard), hide_index=True, width="stretch")

    opt_left, opt_right = st.columns([1.2, 1])
    with opt_left:
        st.subheader("Recommended allocation")
        st.altair_chart(
            spend_shift_chart(optimization["allocation"], "Current Spend", "Recommended Spend"),
            use_container_width=True,
        )
        st.dataframe(
            optimization["allocation"][
                ["Channel", "Current Spend", "Recommended Spend", "Spend Shift", "Change %"]
            ],
            hide_index=True,
            width="stretch",
            column_config={
                "Current Spend": st.column_config.NumberColumn(format="$%.0f"),
                "Recommended Spend": st.column_config.NumberColumn(format="$%.0f"),
                "Spend Shift": st.column_config.NumberColumn(format="$%.0f"),
                "Change %": st.column_config.NumberColumn(format="%.1f%%"),
            },
        )
    with opt_right:
        st.subheader("AI recommendation panel")
        ai_risk_audit = build_responsible_ai_audit(data, evaluation_results, contribution_df, max_mape_pct)
        grounded_payload = build_grounded_ai_payload(
            contribution_df,
            optimization,
            simulation,
            optimization_evidence_packet,
            optimization_scorecard,
            trust_assessment,
            readiness,
            ai_risk_audit,
            optimization_summary,
        )
        narrative = (
            maybe_generate_openai_recommendations(grounded_payload)
            if use_openai
            else None
        )
        deterministic_recommendations = generate_recommendations(contribution_df, optimization, simulation)
        if narrative:
            st.caption("Generated from the MMM evidence packet, KPI scorecard, confidence range, and risk audit.")
            st.markdown(narrative)
        else:
            st.markdown(grounded_ai_brief_html(grounded_payload), unsafe_allow_html=True)
            for item in deterministic_recommendations:
                st.markdown(f"<div class='recommendation'>{item}</div>", unsafe_allow_html=True)

        st.divider()
        st.download_button(
            "Download executive PDF",
            data=build_executive_report_pdf(
                data=data,
                contribution=contribution_df,
                optimization=optimization,
                simulation=simulation,
                evaluation=evaluation_results,
                recommendations=deterministic_recommendations,
                scorecard=optimization_scorecard,
                max_mape_pct=max_mape_pct,
            ),
            file_name="marketing_mix_executive_report.pdf",
            mime="application/pdf",
            key="download_optimization_executive_report",
            width="stretch",
        )
        st.download_button(
            "Download allocation workbook",
            data=build_allocation_workbook(
                optimization,
                optimization_scorecard,
                deterministic_recommendations,
                pilot_plan=build_pilot_plan(optimization, optimization_scorecard, evaluation_results, max_mape_pct),
                risk_audit=build_responsible_ai_audit(data, evaluation_results, contribution_df, max_mape_pct),
            ),
            file_name="recommended_budget_allocation.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_optimization_allocation_workbook",
            width="stretch",
        )
        st.download_button(
            "Download evidence workbook",
            data=build_evidence_workbook(optimization_evidence_packet),
            file_name="mixalyzer_genai_evidence_packet.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_optimization_evidence_packet",
            width="stretch",
        )

with evaluation_tab:
    st.subheader("Model comparison")
    render_badge(str(trust_assessment["label"]), str(trust_assessment["status"]))
    st.caption(
        f"{trust_assessment['explanation']} This matters because budget recommendations should only drive business decisions when the MMM improves on a simple baseline."
    )
    if not model_comparison_df.empty:
        best_model = model_comparison_df.iloc[0]
        cmp_cols = st.columns(4)
        with cmp_cols[0]:
            render_metric_card("Best model", best_model["Model"])
        with cmp_cols[1]:
            render_metric_card("Best MAPE", pct(best_model["MAPE"]))
        with cmp_cols[2]:
            render_metric_card("Selected engine", model.model_kind)
        with cmp_cols[3]:
            render_metric_card("Confidence mode", f"{int(confidence_level * 100)}% interval")
        left_cmp, right_cmp = st.columns([1, 1.25])
        with left_cmp:
            st.altair_chart(model_comparison_chart(model_comparison_df), use_container_width=True)
        with right_cmp:
            st.dataframe(
                model_comparison_df,
                hide_index=True,
                width="stretch",
                column_config={
                    "R2": st.column_config.NumberColumn(format="%.2f"),
                    "MAE": st.column_config.NumberColumn(format="$%.0f"),
                    "RMSE": st.column_config.NumberColumn(format="$%.0f"),
                    "MAPE": st.column_config.NumberColumn(format="%.2f%%"),
                },
            )
    else:
        st.info("Upload at least 12 dated observations to compare model candidates.")

    st.divider()
    st.subheader("Train/test evaluation")
    if evaluation_results:
        eval_cols = st.columns(5)
        evaluation_metrics = [
            ("Train rows", f"{evaluation_results['train_rows']}"),
            ("Test rows", f"{evaluation_results['test_rows']}"),
            ("MMX MAPE", pct(evaluation_results["model_metrics"]["mape"])),
            ("Baseline MAPE", pct(evaluation_results["baseline_metrics"]["mape"])),
            ("RMSE lift", pct(evaluation_results["rmse_improvement_pct"])),
        ]
        for col, (label, value) in zip(eval_cols, evaluation_metrics):
            with col:
                render_metric_card(label, value)

        pred_left, pred_right = st.columns([1.35, 1])
        with pred_left:
            st.altair_chart(prediction_chart(evaluation_results["predictions"]), use_container_width=True)
        with pred_right:
            st.altair_chart(evaluation_metric_chart(evaluation_results), use_container_width=True)

        metrics_df = pd.DataFrame(
            [
                {"Metric": "R-squared", "MMX": evaluation_results["model_metrics"]["r2"], "Baseline": evaluation_results["baseline_metrics"]["r2"]},
                {"Metric": "MAE", "MMX": evaluation_results["model_metrics"]["mae"], "Baseline": evaluation_results["baseline_metrics"]["mae"]},
                {"Metric": "RMSE", "MMX": evaluation_results["model_metrics"]["rmse"], "Baseline": evaluation_results["baseline_metrics"]["rmse"]},
                {"Metric": "MAPE", "MMX": evaluation_results["model_metrics"]["mape"], "Baseline": evaluation_results["baseline_metrics"]["mape"]},
            ]
        )
        st.dataframe(
            metrics_df,
            hide_index=True,
            width="stretch",
            column_config={
                "MMX": st.column_config.NumberColumn(format="%.2f"),
                "Baseline": st.column_config.NumberColumn(format="%.2f"),
            },
        )
    else:
        st.info("Upload at least 12 dated observations to show train/test evaluation.")

with responsible_ai_tab:
    st.subheader("What are the risks?")
    st.caption("This audit frames Mixalyzer as decision support: recommendations should be reviewed before budget changes are activated.")
    risk_df = build_responsible_ai_audit(data, evaluation_results, contribution_df, max_mape_pct)
    passed = int((risk_df["Status"] == "Passed").sum())
    warnings = int((risk_df["Status"] == "Warning").sum())
    needs_review = int((risk_df["Status"] == "Needs review").sum())
    risk_cols = st.columns(4)
    risk_metrics = [
        ("Passed", passed),
        ("Warnings", warnings),
        ("Needs review", needs_review),
        ("Privacy posture", "Aggregated"),
    ]
    for col, (label, value) in zip(risk_cols, risk_metrics):
        with col:
            render_metric_card(label, value)

    for _, row in risk_df.iterrows():
        css_class = "risk-ok" if row["Status"] == "Passed" else "risk-watch"
        st.markdown(
            f"""
            <div class="feature-card {css_class}">
              <h4>{row['Risk Area']} · {row['Status']}</h4>
              <p><strong>Example:</strong> {row['Example in Mixalyzer']}</p>
              <p><strong>Impact:</strong> {row['Potential Impact']}</p>
              <p><strong>Mitigation:</strong> {row['Mitigation / Guardrail']}</p>
            </div>
            """,
            unsafe_allow_html=True,
        )

    st.dataframe(risk_df, hide_index=True, width="stretch")

with pilot_plan_tab:
    st.subheader("Rollout / Pilot Plan")
    st.caption("Use this plan to turn the model recommendation into a controlled business experiment.")
    render_executive_summary(default_summary)

    pilot_duration = st.slider("Pilot duration", min_value=2, max_value=12, value=4, step=1, format="%d weeks")
    pilot_plan = build_pilot_plan(
        default_optimization,
        business_scorecard,
        evaluation_results,
        max_mape_pct,
        duration_weeks=pilot_duration,
    )
    st.dataframe(pilot_plan, hide_index=True, width="stretch")

    shifts = allocation_shift_summary(default_optimization)
    pilot_cols = st.columns(3)
    with pilot_cols[0]:
        render_metric_card("Channels to increase", shifts["increase"])
    with pilot_cols[1]:
        render_metric_card("Channels to decrease", shifts["decrease"])
    with pilot_cols[2]:
        render_metric_card("Monitoring cadence", "Weekly")

    st.markdown(
        """
        <div class="summary-box">
          <h4>Stop-loss and review rules</h4>
          <p>Pause or revise if conservative lift becomes negative for 2 consecutive weeks.</p>
          <p>Pause if CAC increases beyond the selected target or conversion volume drops materially.</p>
          <p>Pause if model error exceeds the acceptable MAPE threshold.</p>
          <p>Require a human review checkpoint before permanent rollout.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

with model_tab:
    model_left, model_right = st.columns([1, 1])
    with model_left:
        st.subheader("Training data")
        st.dataframe(data.tail(12), hide_index=True, width="stretch")
        st.subheader("Adstock carryover settings")
        adstock_df = pd.DataFrame(
            [
                {
                    "Channel": CHANNEL_LABELS[channel],
                    "Carryover Decay": DEFAULT_ADSTOCK_DECAYS[channel],
                    "Interpretation": "Higher values mean spend influence lasts longer.",
                }
                for channel in DEFAULT_CHANNELS
            ]
        )
        st.dataframe(
            adstock_df,
            hide_index=True,
            width="stretch",
            column_config={"Carryover Decay": st.column_config.NumberColumn(format="%.2f")},
        )
        st.subheader("External controls included")
        st.dataframe(detected_control_summary(data), hide_index=True, width="stretch")
    with model_right:
        st.subheader("Model diagnostics")
        diagnostics = pd.DataFrame(
            [
                {"Metric": "R-squared", "Value": model.metrics["r2"]},
                {"Metric": "MAE", "Value": model.metrics["mae"]},
                {"Metric": "RMSE", "Value": model.metrics["rmse"]},
                {"Metric": "MAPE", "Value": model.metrics["mape"]},
            ]
        )
        st.dataframe(diagnostics, hide_index=True, width="stretch")

        coefficients = (
            model.coefficients.rename_axis("Feature")
            .reset_index(name="Coefficient")
            .sort_values("Coefficient", ascending=False)
        )
        st.dataframe(coefficients, hide_index=True, width="stretch")
